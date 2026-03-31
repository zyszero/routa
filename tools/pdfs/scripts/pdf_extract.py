#!/usr/bin/env python3
"""Extract text/layout/tables/images/forms/attachments from PDFs.

Subcommands:
  info         - quick summary (delegates to pdf_inspect)
  text         - extract plain text
  words        - extract word boxes (coordinates)
  chars        - extract character boxes (fine-grained layout)
  tables       - extract tables to CSV or XLSX
  images       - extract embedded images
  attachments  - extract embedded file attachments
  annotations  - list annotations
  forms        - list form fields

Examples:
  python pdf_extract.py text input.pdf --method pdfplumber --out /mnt/data/text.txt
  python pdf_extract.py words input.pdf --out words.csv
  python pdf_extract.py tables input.pdf --out_dir /mnt/data/tables --format xlsx
  python pdf_extract.py images input.pdf --out_dir /mnt/data/images

Notes:
  - pdfplumber coordinates are top-left origin; y increases downward.
  - PyMuPDF (fitz) word boxes are also top-left origin.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
from dataclasses import asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


def _parse_page_ranges(spec: Optional[str]) -> Optional[List[Tuple[int, int]]]:
    if not spec:
        return None
    spec = spec.strip()
    if not spec:
        return None
    ranges: List[Tuple[int, int]] = []
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            a_i = int(a)
            b_i = int(b)
            if a_i <= 0 or b_i <= 0:
                raise ValueError("Page numbers must be >= 1")
            if b_i < a_i:
                a_i, b_i = b_i, a_i
            ranges.append((a_i, b_i))
        else:
            p = int(part)
            if p <= 0:
                raise ValueError("Page numbers must be >= 1")
            ranges.append((p, p))
    return ranges


def _iter_pages(num_pages: int, ranges: Optional[List[Tuple[int, int]]]) -> Iterable[int]:
    if not ranges:
        yield from range(1, num_pages + 1)
        return
    for a, b in ranges:
        a = max(1, a)
        b = min(num_pages, b)
        for p in range(a, b + 1):
            yield p


def _ensure_dir(path: str) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _read_pdf_text_pdftotext(input_pdf: str, pages: Optional[List[Tuple[int, int]]]) -> str:
    # pdftotext does not support disjoint ranges directly, so we extract all and slice at paragraph boundaries.
    # For large files, prefer python methods (pdfplumber/pymupdf) with page control.
    cmd = ["pdftotext", input_pdf, "-"]
    proc = subprocess.run(cmd, check=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.decode("utf-8", "ignore").strip() or "pdftotext failed")
    text = proc.stdout.decode("utf-8", "ignore")
    if not pages:
        return text
    # Best-effort: keep full text (page boundaries are not reliably preserved).
    return text


def cmd_text(args: argparse.Namespace) -> int:
    method = args.method
    ranges = _parse_page_ranges(args.pages)

    if method == "pdftotext":
        text = _read_pdf_text_pdftotext(args.input_pdf, ranges)
    elif method == "pypdf":
        from pypdf import PdfReader
        from pypdf.generic import DictionaryObject, StreamObject

        reader = PdfReader(args.input_pdf)
        out_parts: List[str] = []
        for p in _iter_pages(len(reader.pages), ranges):
            out_parts.append(reader.pages[p - 1].extract_text() or "")
        text = "\n\n".join(out_parts)
    elif method == "pdfplumber":
        import pdfplumber

        out_parts = []
        with pdfplumber.open(args.input_pdf) as pdf:
            for p in _iter_pages(len(pdf.pages), ranges):
                out_parts.append(pdf.pages[p - 1].extract_text() or "")
        text = "\n\n".join(out_parts)
    elif method == "pymupdf":
        import fitz

        doc = fitz.open(args.input_pdf)
        out_parts = []
        for p in _iter_pages(doc.page_count, ranges):
            page = doc.load_page(p - 1)
            out_parts.append(page.get_text("text") or "")
        text = "\n\n".join(out_parts)
    else:
        raise ValueError(f"Unknown method: {method}")

    if args.out:
        Path(args.out).write_text(text, encoding="utf-8")
    else:
        print(text)
    return 0


def cmd_words(args: argparse.Namespace) -> int:
    ranges = _parse_page_ranges(args.pages)

    rows: List[Dict[str, Any]] = []

    if args.method == "pdfplumber":
        import pdfplumber

        with pdfplumber.open(args.input_pdf) as pdf:
            for p in _iter_pages(len(pdf.pages), ranges):
                page = pdf.pages[p - 1]
                for w in page.extract_words():
                    rows.append(
                        {
                            "page": p,
                            "text": w.get("text", ""),
                            "x0": w.get("x0"),
                            "x1": w.get("x1"),
                            "top": w.get("top"),
                            "bottom": w.get("bottom"),
                            "doctop": w.get("doctop"),
                        }
                    )
    elif args.method == "pymupdf":
        import fitz

        doc = fitz.open(args.input_pdf)
        for p in _iter_pages(doc.page_count, ranges):
            page = doc.load_page(p - 1)
            for x0, y0, x1, y1, word, block_no, line_no, word_no in page.get_text("words"):
                rows.append(
                    {
                        "page": p,
                        "text": word,
                        "x0": x0,
                        "y0": y0,
                        "x1": x1,
                        "y1": y1,
                        "block": block_no,
                        "line": line_no,
                        "word_no": word_no,
                    }
                )
    else:
        raise ValueError("--method must be pdfplumber or pymupdf")

    if not args.out:
        print(json.dumps(rows, indent=2, ensure_ascii=False))
        return 0

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.suffix.lower() == ".json":
        out_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
        return 0

    # CSV
    fieldnames = sorted({k for r in rows for k in r.keys()})
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return 0




def cmd_chars(args: argparse.Namespace) -> int:
    """Extract character-level boxes.

    Uses pdfplumber only (reliable char metrics). Output is CSV or JSON depending
    on --out extension. Coordinates follow pdfplumber convention: origin top-left,
    y increases downward.
    """

    import pdfplumber

    ranges = _parse_page_ranges(args.pages)
    rows = []

    with pdfplumber.open(args.input_pdf) as pdf:
        for p in _iter_pages(len(pdf.pages), ranges):
            page = pdf.pages[p - 1]
            for ch in (page.chars or []):
                rows.append({
                    'page': p,
                    'text': ch.get('text', ''),
                    'x0': ch.get('x0'),
                    'x1': ch.get('x1'),
                    'top': ch.get('top'),
                    'bottom': ch.get('bottom'),
                    'doctop': ch.get('doctop'),
                    'fontname': ch.get('fontname'),
                    'size': ch.get('size'),
                })

    if not args.out:
        print(json.dumps(rows, indent=2, ensure_ascii=False))
        return 0

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.suffix.lower() == '.json':
        out_path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding='utf-8')
        return 0

    fieldnames = sorted({k for r in rows for k in r.keys()})
    with out_path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return 0

def cmd_tables(args: argparse.Namespace) -> int:
    import pdfplumber

    ranges = _parse_page_ranges(args.pages)
    out_dir = _ensure_dir(args.out_dir)

    tables_for_xlsx: List[Tuple[str, List[List[Optional[str]]]]] = []

    with pdfplumber.open(args.input_pdf) as pdf:
        for p in _iter_pages(len(pdf.pages), ranges):
            page = pdf.pages[p - 1]
            tables = page.extract_tables()
            for ti, table in enumerate(tables, start=1):
                # Normalize to strings
                norm: List[List[Optional[str]]] = []
                for row in table:
                    if row is None:
                        continue
                    norm.append([None if c is None else str(c) for c in row])

                name = f"page{p:03d}_table{ti:02d}"

                if args.format == "csv":
                    csv_path = out_dir / f"{name}.csv"
                    with csv_path.open("w", newline="", encoding="utf-8") as f:
                        writer = csv.writer(f)
                        for row in norm:
                            writer.writerow(row)
                elif args.format == "json":
                    json_path = out_dir / f"{name}.json"
                    json_path.write_text(json.dumps(norm, indent=2, ensure_ascii=False), encoding="utf-8")
                elif args.format == "xlsx":
                    tables_for_xlsx.append((name, norm))
                else:
                    raise ValueError("--format must be csv|json|xlsx")

    if args.format == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception as e:  # pragma: no cover
            raise RuntimeError("openpyxl is required for xlsx output") from e

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for sheet_name, table in tables_for_xlsx:
            ws = wb.create_sheet(title=sheet_name[:31])
            for r_idx, row in enumerate(table, start=1):
                for c_idx, cell in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=cell)

        xlsx_path = out_dir / "tables.xlsx"
        wb.save(xlsx_path)

    print(f"Wrote tables to: {out_dir}")
    return 0


def cmd_images(args: argparse.Namespace) -> int:
    out_dir = _ensure_dir(args.out_dir)
    pages = _parse_page_ranges(args.pages)

    if args.method == "pdfimages":
        # Uses poppler to dump images
        cmd = ["pdfimages", "-all", args.input_pdf, str(out_dir / "img")]
        proc = subprocess.run(cmd, check=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if proc.returncode != 0:
            raise RuntimeError(proc.stderr.strip() or "pdfimages failed")
        print(f"Extracted images to: {out_dir}")
        return 0

    if args.method != "pymupdf":
        raise ValueError("--method must be pymupdf or pdfimages")

    import fitz

    doc = fitz.open(args.input_pdf)
    count = 0
    for p in _iter_pages(doc.page_count, pages):
        page = doc.load_page(p - 1)
        for img_i, img in enumerate(page.get_images(full=True), start=1):
            xref = img[0]
            pix = fitz.Pixmap(doc, xref)
            if pix.n - pix.alpha > 3:
                pix = fitz.Pixmap(fitz.csRGB, pix)
            out_path = out_dir / f"page{p:03d}_img{img_i:03d}.png"
            pix.save(out_path)
            count += 1
    print(f"Extracted {count} images to: {out_dir}")
    return 0


def cmd_attachments(args: argparse.Namespace) -> int:
    from pypdf import PdfReader

    out_dir = _ensure_dir(args.out_dir)
    reader = PdfReader(args.input_pdf)
    attachments = getattr(reader, "attachments", {}) or {}

    written = 0
    for name, payload in attachments.items():
        # pypdf may return a single bytes or a list of bytes for the same name
        if isinstance(payload, list):
            items = payload
        else:
            items = [payload]

        for idx, data in enumerate(items, start=1):
            safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", name) or "attachment"
            if len(items) > 1:
                safe_name = f"{safe_name}.{idx}"
            out_path = out_dir / safe_name
            out_path.write_bytes(data)
            written += 1

    print(f"Wrote {written} attachments to: {out_dir}")
    return 0


def cmd_annotations(args: argparse.Namespace) -> int:
    from pypdf import PdfReader

    reader = PdfReader(args.input_pdf)
    pages = _parse_page_ranges(args.pages)

    annots_out: List[Dict[str, Any]] = []

    for p in _iter_pages(len(reader.pages), pages):
        page = reader.pages[p - 1]
        annots = page.get("/Annots")
        if not annots:
            continue
        for a in annots:
            try:
                obj = a.get_object()
            except Exception:
                continue
            annots_out.append(
                {
                    "page": p,
                    "subtype": str(obj.get("/Subtype")),
                    "rect": [float(x) for x in (obj.get("/Rect") or [])],
                    "contents": str(obj.get("/Contents")) if obj.get("/Contents") is not None else None,
                    "name": str(obj.get("/NM")) if obj.get("/NM") is not None else None,
                }
            )

    if args.out:
        Path(args.out).write_text(json.dumps(annots_out, indent=2, ensure_ascii=False), encoding="utf-8")
    else:
        print(json.dumps(annots_out, indent=2, ensure_ascii=False))
    return 0


def cmd_forms(args: argparse.Namespace) -> int:
    from pypdf import PdfReader
    from pypdf.generic import DictionaryObject, StreamObject

    reader = PdfReader(args.input_pdf)
    fields = reader.get_fields() or {}

    def _obj(o: Any) -> Any:
        try:
            return o.get_object()
        except Exception:
            return o

    def _name_for_field(field_obj: Any) -> str:
        """Best-effort fully-qualified name.

        PDF field names can be hierarchical via /Parent.
        A common convention is joining /T components with '.'
        (viewers vary; this is best-effort for debugging).
        """

        parts: List[str] = []
        cur = field_obj
        visited = set()
        while cur is not None:
            cur = _obj(cur)
            oid = getattr(cur, "idnum", None)
            if oid is not None and oid in visited:
                break
            if oid is not None:
                visited.add(oid)

            t = cur.get("/T") if hasattr(cur, "get") else None
            if t is not None:
                parts.append(str(t))
            cur = cur.get("/Parent") if hasattr(cur, "get") else None
        parts = list(reversed(parts))
        return ".".join(parts) if parts else ""

    out: Dict[str, Any] = {}
    for name, field in fields.items():
        f = _obj(field)
        info: Dict[str, Any] = {
            "type": str(f.get("/FT")) if hasattr(f, "get") else None,
            "value": str(f.get("/V")) if hasattr(f, "get") else None,
            "default": str(f.get("/DV")) if hasattr(f, "get") else None,
            "alt_name": str(f.get("/TU")) if hasattr(f, "get") else None,
            "flags": int(f.get("/Ff")) if hasattr(f, "get") and f.get("/Ff") is not None else None,
        }

        # Choice fields: dropdown/list options
        if hasattr(f, "get") and f.get("/Opt") is not None:
            try:
                opt = f.get("/Opt")
                # /Opt can be ["A","B"] or [["export","display"], ...]
                info["options"] = [
                    (list(x) if isinstance(x, (list, tuple)) else str(x)) for x in opt
                ]
            except Exception:
                info["options"] = None

        out[name] = info

    if args.include_widgets:
        # Collect widget annotations (page + rect) to debug alignment issues.
        for pageno, page in enumerate(reader.pages, start=1):
            annots = page.get("/Annots")
            if not annots:
                continue
            for a in annots:
                annot = _obj(a)
                if not hasattr(annot, "get"):
                    continue
                if str(annot.get("/Subtype")) != "/Widget":
                    continue

                # Determine field dictionary: widget itself, or /Parent chain.
                field_obj = annot
                if annot.get("/Parent") is not None:
                    field_obj = _obj(annot.get("/Parent"))

                # Prefer the name from the widget's /T if present, else parent chain.
                widget_name = _name_for_field(annot) or _name_for_field(field_obj)

                # Fall back to pypdf's get_fields keys by matching suffix.
                chosen_name = widget_name
                if not chosen_name:
                    chosen_name = str(annot.get("/T")) if annot.get("/T") is not None else ""
                if chosen_name not in out and chosen_name:
                    # Try to match on ending component
                    for k in out.keys():
                        if k.endswith(chosen_name):
                            chosen_name = k
                            break

                rect = annot.get("/Rect") or []
                rect_f = [float(x) for x in rect] if rect else None

                # Appearance states (useful for checkboxes/radios).
                # IMPORTANT: /AP /N can be either:
                #   - a stream (normal appearance content), OR
                #   - a dictionary of named appearances (e.g. /Off, /Yes) for widgets.
                # Streams are also dict-like in pypdf, so we must distinguish them.
                states = None
                try:
                    ap = annot.get("/AP")
                    nobj = ap.get("/N") if ap else None
                    if isinstance(nobj, DictionaryObject) and not isinstance(nobj, StreamObject):
                        # Named appearance dictionary -> treat keys as states.
                        states = [str(k) for k in nobj.keys()]
                except Exception:
                    states = None

                widget_info = {
                    "page": pageno,
                    "rect": rect_f,
                    "appearance_states": states,
                }

                if chosen_name:
                    out.setdefault(chosen_name, {})
                    out[chosen_name].setdefault("widgets", [])
                    out[chosen_name]["widgets"].append(widget_info)
                else:
                    out.setdefault("__unmatched_widgets__", []).append(widget_info)

    if args.out:
        Path(args.out).write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    else:
        print(json.dumps(out, indent=2, ensure_ascii=False))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_info = sub.add_parser("info", help="inspect pdf")
    p_info.add_argument("input_pdf")
    p_info.add_argument("--json", action="store_true")
    p_info.add_argument("--password", default=None)

    p_text = sub.add_parser("text", help="extract plain text")
    p_text.add_argument("input_pdf")
    p_text.add_argument("--method", choices=["pdfplumber", "pymupdf", "pypdf", "pdftotext"], default="pdfplumber")
    p_text.add_argument("--pages", default=None, help="e.g. 1-3,5")
    p_text.add_argument("--out", "--output", dest="out", default=None)

    p_words = sub.add_parser("words", help="extract word boxes")
    p_words.add_argument("input_pdf")
    p_words.add_argument("--method", choices=["pdfplumber", "pymupdf"], default="pdfplumber")
    p_words.add_argument("--pages", default=None)
    p_words.add_argument("--out", "--output", dest="out", default=None, help=".csv or .json; if omitted prints JSON")

    p_chars = sub.add_parser("chars", help="extract character boxes")
    p_chars.add_argument("input_pdf")
    p_chars.add_argument("--pages", default=None)
    p_chars.add_argument("--out", "--output", dest="out", default=None, help=".csv or .json; if omitted prints JSON")

    p_tables = sub.add_parser("tables", help="extract tables")
    p_tables.add_argument("input_pdf")
    p_tables.add_argument("--pages", default=None)
    p_tables.add_argument("--out_dir", required=True)
    p_tables.add_argument("--format", choices=["csv", "json", "xlsx"], default="csv")

    p_images = sub.add_parser("images", help="extract images")
    p_images.add_argument("input_pdf")
    p_images.add_argument("--pages", default=None)
    p_images.add_argument("--out_dir", required=True)
    p_images.add_argument("--method", choices=["pymupdf", "pdfimages"], default="pymupdf")

    p_att = sub.add_parser("attachments", help="extract embedded attachments")
    p_att.add_argument("input_pdf")
    p_att.add_argument("--out_dir", required=True)

    p_ann = sub.add_parser("annotations", help="list annotations")
    p_ann.add_argument("input_pdf")
    p_ann.add_argument("--pages", default=None)
    p_ann.add_argument("--out", "--output", dest="out", default=None)

    p_forms = sub.add_parser("forms", help="list form fields")
    p_forms.add_argument("input_pdf")
    p_forms.add_argument("--out", "--output", dest="out", default=None)
    p_forms.add_argument("--include_widgets", "--include-widgets", dest="include_widgets", action="store_true")

    args = parser.parse_args()

    if args.cmd == "info":
        from pdf_inspect import _print_human, inspect_pdf

        input_pdf = Path(args.input_pdf)
        if not input_pdf.exists():
            print(f"ERROR: not found: {input_pdf}", file=sys.stderr)
            return 2
        summary = inspect_pdf(input_pdf, password=args.password)
        if args.json:
            print(json.dumps(asdict(summary), indent=2, ensure_ascii=True))
        else:
            _print_human(summary)
        return 0

    if args.cmd == "text":
        return cmd_text(args)
    if args.cmd == "words":
        return cmd_words(args)
    if args.cmd == "chars":
        return cmd_chars(args)
    if args.cmd == "tables":
        return cmd_tables(args)
    if args.cmd == "images":
        return cmd_images(args)
    if args.cmd == "attachments":
        return cmd_attachments(args)
    if args.cmd == "annotations":
        return cmd_annotations(args)
    if args.cmd == "forms":
        return cmd_forms(args)

    raise AssertionError(f"Unhandled cmd: {args.cmd}")


if __name__ == "__main__":
    raise SystemExit(main())
